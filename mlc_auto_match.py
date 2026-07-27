"""
mlc_auto_match.py

Automazione Playwright per The MLC Matching Tool.

Il processo di match su MLC ha DUE stage distinti e sequenziali, non un
singolo step con fallback:

  Stage 1 - Ricerca della REGISTRAZIONE (recording) per ISRC.
            Se trova piu' gruppi per lo stesso ISRC (es. varianti "Original
            Mix" arrivate da DSP diversi), si selezionano TUTTI e si
            confermano insieme - non e' un caso ambiguo, e' normale. Se
            l'ISRC non trova nulla, la riga e' no_match_recording. Se il
            gruppo risulta gia' "Submitted"/"Accepted"/"Rejected" da una
            sessione precedente, la riga e' already_submitted e si passa
            all'ISRC successivo senza toccare nulla.

  Stage 2 - Ricerca della TUA OPERA (work) gia' registrata nel catalogo, da
            abbinare alla registrazione confermata allo Stage 1. Si cerca per
            titolo + un secondo criterio: di default Publisher Name con il
            valore fisso "LOO" (funziona per "LOOSE CLUB EDITION" a
            prescindere dal publisher esatto in riga). Se questo non trova
            nulla o e' ambiguo (piu' di un risultato) e il writer e'
            disponibile in input, si ritenta con Titolo + Writer Name
            (cognome autore), piu' selettivo sui titoli generici tipo
            "System" o "Contacto" dove il publisher da solo non discrimina.

Le righe ancora ambigue dopo il fallback (piu' risultati su entrambi i
criteri, o layout inatteso) vengono messe in pausa per conferma manuale
invece di essere decise a caso. Se lo Stage 2 non trova nulla nemmeno col
fallback, l'opera probabilmente non e' ancora registrata nel catalogo: la
riga viene segnata no_match_work per verifica manuale, non e' un errore
dello script.

Setup:
    pip install -r requirements.txt
    npx playwright install chromium

    export MLC_EMAIL="tua_email@dominio.com"
    export MLC_PASSWORD="tua_password"

Uso:
    python mlc_auto_match.py catalogo.xlsx --sheet "FEB 25" --output risultati.xlsx
"""

import argparse
import os
import re
import sys
import time
import unicodedata
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from playwright.sync_api import sync_playwright, Page, TimeoutError as PWTimeout

LOGIN_URL = "https://portal.themlc.com/login"

# testi di "nessun risultato" osservati nei due stage (sono formulazioni diverse)
NO_RESULTS_PATTERNS = re.compile(
    r"couldn.?t locate any results|no results found", re.IGNORECASE
)

# colorazione riga nel report Excel, come nel flusso manuale: verde quando
# c'e' un match (o e' gia' stato gestito prima), giallo quando non si trova
# nulla e serve verifica manuale
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
# arancione: non e' "opera non trovata" (giallo) ma un tentativo fallito da
# ritentare - colorarlo di giallo lo confonderebbe con le opere non a catalogo
ORANGE_FILL = PatternFill(start_color="FFC7A0", end_color="FFC7A0", fill_type="solid")
GREEN_STATUSES = {"matched", "already_submitted"}
YELLOW_STATUSES = {"no_match_recording", "no_match_work"}
ORANGE_STATUSES = {"submit_failed", "manual_incomplete", "error"}


@dataclass
class RowResult:
    isrc: str
    title: str
    writer: str
    publisher: str
    status: str = "pending"
    # matched / no_match_recording / no_match_work / already_submitted /
    # ambiguous_recording / ambiguous_work / submit_failed /
    # manual_incomplete / error
    note: str = ""


COOKIE_ACCEPT_ID = "#CybotCookiebotDialogBodyLevelButtonLevelOptinAllowAll"

# limite di sicurezza per i cicli di selezione/deselezione dei gruppi, per
# non rischiare loop infiniti se il DOM non si aggiorna come previsto
MAX_GROUPS_PER_ISRC = 50


def _dismiss_cookies(page: Page) -> None:
    """
    Il banner cookie (widget Cookiebot) compare su ogni browser senza
    cookie salvati e blocca i click sottostanti se resta a schermo. Elenca
    una decina di vendor di tracking e puo' metterci diversi secondi a
    diventare cliccabile (osservato live: un tentativo di 3s falliva e il
    dialog restava, facendo scadere il click su "Login" dopo 30s) - qui si
    aspetta fino a 10s e si clicca l'id stabile del pulsante "Allow all
    cookies", senza bloccare se il banner non compare affatto.
    """
    accept_button = page.locator(COOKIE_ACCEPT_ID)
    try:
        accept_button.click(timeout=10000)
    except PWTimeout:
        pass


def login(page: Page, email: str, password: str) -> None:
    page.goto(LOGIN_URL)
    _dismiss_cookies(page)

    page.get_by_role("textbox", name="Email Address").fill(email)
    page.get_by_role("textbox", name="Password").fill(password)

    _dismiss_cookies(page)  # puo' comparire solo a questo punto del flusso

    page.get_by_role("button", name="Login").click()

    # il 2FA non viene richiesto a ogni login (es. dispositivo/sessione gia'
    # riconosciuta dal sito): se il campo Code non compare entro pochi
    # secondi si salta lo step, il login e' gia' andato a buon fine
    # (osservato live: altrimenti lo script resta bloccato in attesa di un
    # campo che non arrivera' mai)
    code_box = page.get_by_role("textbox", name="Code")
    try:
        code_box.wait_for(timeout=8000)
    except PWTimeout:
        code_box = None

    if code_box is not None:
        otp = input("Inserisci il codice OTP ricevuto via email/SMS: ").strip()
        code_box.fill(otp)
        page.get_by_role("button", name="Login").click()

    page.get_by_role("link", name="Matching Tool", exact=True).click()
    page.wait_for_load_state("networkidle")


def reset_search_form(page: Page) -> None:
    """
    Torna alla schermata di ricerca Stage 1 dopo un esito che non ha mai
    selezionato un gruppo recording (no_match_recording, already_submitted)
    o dopo un match/ambiguous_work completato con Confirm+Done. Il click e'
    di fatto un no-op quando si e' gia' su questa route (SPA, nessun
    remount) - il valore del campo ISRC resta finche' non lo si sovrascrive
    con .fill() sulla riga successiva, e va bene cosi'.
    """
    page.get_by_role("link", name="Matching Tool", exact=True).click()
    page.wait_for_load_state("networkidle")


def abandon_stage2(page: Page) -> None:
    """
    Torna da Stage 2 (nessun'opera trovata) a Stage 1 SENZA confermare
    nulla. A differenza di Confirm+Done, "Back" lascia i gruppi recording
    ancora selezionati: vanno deselezionati TUTTI, altrimenti resterebbero
    agganciati alla ricerca ISRC successiva.

    Con piu' gruppi selezionati le icone "deselect" sono altrettante e un
    click singolo solleva "strict mode violation" (verificato live con 3
    gruppi), quindi si cicla sulla prima rimasta finche' non spariscono.
    """
    page.get_by_role("button", name="Back").click()
    page.wait_for_timeout(500)

    for _ in range(MAX_GROUPS_PER_ISRC):
        icons = page.get_by_role("img", name="deselect")
        if icons.count() == 0:
            break
        icons.first.click()
        page.wait_for_timeout(300)


# endpoint di ricerca dei due stage: Stage 1 GET .../search/unmatched-recordings,
# Stage 2 POST .../search/works/catalog
SEARCH_ENDPOINT_RE = re.compile(r"/search/(unmatched-recordings|works/catalog)")

# la pagina ha finito di mostrare l'esito quando compare il messaggio di
# "nessun risultato" oppure almeno un risultato/intestazione della lista
RESULTS_RENDERED_JS = """() => {
  const t = document.body.innerText;
  return /couldn.?t locate any results|no results found/i.test(t)
      || /Showing\\s+\\d+\\s*-\\s*\\d+\\s+of\\s+\\d+\\s+results/i.test(t)
      || document.querySelectorAll('[data-testid="select-rg-button"]').length > 0
      || document.querySelectorAll('#select-link').length > 0;
}"""


def run_search(page: Page) -> str:
    """
    Esegue la ricerca e attende che l'esito sia effettivamente a schermo.
    Ritorna 'no_results' oppure 'has_results'.

    Non si usa un'attesa a tempo fisso: con la pagina lenta (tipicamente la
    prima ricerca dopo il login) i risultati non erano ancora renderizzati,
    il codice contava zero pulsanti di selezione e classificava la riga come
    'ambiguous_recording' pur essendo un caso normalissimo (osservato live
    su ISRC CARH11900303, che rifatto a mano mostra 2 gruppi regolari).
    Si aspetta prima la RISPOSTA DI RETE della ricerca - cosi' non si rischia
    di leggere i risultati della ricerca precedente rimasti a schermo - e poi
    il render vero e proprio.
    """
    try:
        with page.expect_response(
            lambda r: bool(SEARCH_ENDPOINT_RE.search(r.url)), timeout=30000
        ):
            page.get_by_role("button", name="Search").click()
    except PWTimeout:
        page.wait_for_timeout(1200)  # nessuna chiamata intercettata: margine di sicurezza

    try:
        page.wait_for_function(RESULTS_RENDERED_JS, timeout=15000)
    except PWTimeout:
        pass  # si prosegue comunque: la classificazione sotto gestisce anche il caso vuoto
    page.wait_for_timeout(300)

    if page.get_by_text(NO_RESULTS_PATTERNS).count() > 0:
        return "no_results"
    return "has_results"


def _load_all_results(page: Page) -> None:
    """
    Oltre una certa soglia di risultati compare un pulsante 'Load More' che
    puo' ripresentarsi piu' volte (paginazione incrementale): va cliccato
    finche' non scompare, altrimenti si selezionerebbero solo i primi
    risultati caricati e non tutti i gruppi.
    """
    load_more = page.get_by_role("button", name=re.compile(r"load more", re.IGNORECASE))
    while load_more.count() > 0:
        load_more.click()
        page.wait_for_timeout(800)
        load_more = page.get_by_role("button", name=re.compile(r"load more", re.IGNORECASE))


# ---------------------------------------------------------------------------
# Stage 1 - ricerca della registrazione (recording) per ISRC
# ---------------------------------------------------------------------------

ALREADY_PROCESSED_PATTERN = re.compile(r"Submitted|Accepted|Rejected")


def _ensure_isrc_criteria(page: Page) -> None:
    """
    La riga 1 ('search.1.searchTerm') e' lo STESSO slot di criterio
    riusato anche dallo Stage 2 (stesso test-id), non un campo dedicato:
    se una ricerca precedente e' finita in errore a meta' di un cambio
    criterio in Stage 2, questo slot puo' restare "sporco" (es. bloccato su
    "MLC Song Code" invece di "Recording ISRC") e l'ISRC finirebbe cercato
    nel campo sbagliato senza che lo script se ne accorga - niente eccezioni,
    solo risultati falsi (osservato live su un'intera run dopo un errore).
    Si verifica che l'etichetta sia davvero "Recording ISRC" e, se non lo
    e', si ricarica la pagina da zero per ripristinare lo stato genuino
    (un click sul link "Matching Tool" non basta, e' un no-op SPA).
    """
    if page.get_by_text("Recording ISRC", exact=True).count() == 0:
        page.reload(wait_until="networkidle")
        _dismiss_cookies(page)  # il banner puo' ricomparire dopo un reload e bloccare i click
        page.get_by_role("link", name="Matching Tool", exact=True).click()
        page.wait_for_load_state("networkidle")


def _select_all_groups(page: Page, count: int) -> None:
    """
    Seleziona tutti i gruppi trovati per l'ISRC.

    ATTENZIONE agli indici: quando un gruppo viene selezionato il suo
    'select-rg-button' SPARISCE (diventa l'icona "deselect"), quindi la lista
    si accorcia a ogni click e gli indici slittano. Iterare con nth(i) fisso
    salta un gruppo su due e poi va in timeout (osservato live: 3 gruppi
    trovati, selezionati solo il 1o e il 3o -> "Match 2 Groups", poi errore
    su nth(2) inesistente). Si clicca sempre il PRIMO rimasto, tante volte
    quanti sono i gruppi.
    """
    for _ in range(count):
        buttons = page.get_by_test_id("select-rg-button")
        if buttons.count() == 0:
            break
        buttons.first.click()
        page.wait_for_timeout(300)


def stage1_search_isrc(page: Page, isrc: str) -> str:
    """
    Il form di ricerca recording ha di default due righe gia' presenti:
    riga 0 = Recording Title, riga 1 = Recording ISRC. Non serve toccare
    il dropdown dei criteri, si compila direttamente la riga 1.
    Ritorna: 'matched' / 'no_match' / 'already_submitted' / 'ambiguous'
    """
    _ensure_isrc_criteria(page)

    isrc_box = page.get_by_test_id("search.1.searchTerm")
    isrc_box.click()
    isrc_box.fill(isrc)

    status = run_search(page)
    if status == "no_results":
        return "no_match"

    _load_all_results(page)  # 'Load More' puo' comparire piu' volte: va esaurito prima di selezionare

    select_rg = page.get_by_test_id("select-rg-button")
    count = select_rg.count()

    if count == 0:
        # nessun controllo di selezione: o e' gia' stato processato in una
        # sessione precedente (niente da fare, si salta), o e' un layout
        # inatteso rispetto a quanto osservato
        if page.get_by_text(ALREADY_PROCESSED_PATTERN).count() > 0:
            return "already_submitted"
        return "ambiguous"

    # piu' gruppi trovati per lo stesso ISRC (es. varianti "Original Mix" da
    # DSP diversi): si selezionano TUTTI e si matchano insieme, non e' un
    # caso ambiguo da mettere in pausa
    _select_all_groups(page, count)

    match_button = page.get_by_role("button", name=re.compile(r"^Match \d+ Groups?$"))
    if match_button.count() == 0:
        return "ambiguous"

    match_button.click()
    page.get_by_role("button", name="Continue").click()
    return "matched"


# ---------------------------------------------------------------------------
# Stage 2 - ricerca dell'opera (work) da abbinare alla registrazione confermata
# ---------------------------------------------------------------------------

CRITERIA_LABELS = [
    "MLC Song Code",
    "Publisher Name",
    "Writer Name",
    "ISWC",
    "Writer IPI",
    "Publisher IPI",
    "MLC Publisher Number",
]


def _open_criteria_menu(page: Page, container) -> None:
    """
    Apre il menu a tendina dei criteri.

    Il div con data-testid copre TUTTA la riga (etichetta + casella di testo
    + toggle "Exact Match"): cliccarlo al centro finisce sulla CASELLA DI
    TESTO, non sulla tendina, quindi il menu non si apre mai e la successiva
    attesa dell'opzione va in timeout dopo 30s (osservato live: etichetta
    rimasta su "MLC Song Code" e cursore lampeggiante nel campo vuoto).
    Va cliccata l'ETICHETTA con il criterio corrente, come faceva la
    sessione Codegen originale.
    """
    for label in CRITERIA_LABELS:
        target = container.get_by_text(label, exact=True)
        if target.count() > 0:
            target.first.click()
            return
    container.click()  # fallback: etichetta inattesa, meglio provare che fallire


def _set_second_criteria(page: Page, criteria_name: str, value: str) -> None:
    """
    Nello Stage 2 la riga 0 (Work Title) e' gia' presente di default. La riga 1
    (secondo criterio) di norma esiste gia' anch'essa - non va aggiunta con
    'Add Criteria': si clicca direttamente il suo dropdown, che mantiene il
    valore scelto nella ricerca precedente finche' non lo si cambia (osservato
    via Codegen su piu' righe). Se il form e' stato resettato e la riga non
    c'e' ancora, si ricade su 'Add Criteria'.
    """
    dropdown = page.locator('[data-testid^="search-with-criteria-select-div-"]')
    if dropdown.count() == 0:
        page.get_by_role("button", name="Add Criteria").click()
        dropdown = page.locator('[data-testid^="search-with-criteria-select-div-"]')
    _open_criteria_menu(page, dropdown.last)
    page.get_by_role("option", name=criteria_name).click()

    term_box = page.get_by_test_id("search.1.searchTerm")
    term_box.click()
    term_box.fill(value)


def _count_work_results(page: Page) -> int:
    # ogni riga risultato nello Stage 2 ha un link con id 'select-link'
    # (id ripetuto su ogni riga, non un pulsante - osservato via Codegen)
    return page.locator("#select-link").count()


# estrae il testo di ogni scheda risultato dello Stage 2: si risale dal
# 'select-link' fino al primo antenato comune a tutti i risultati, la scheda
# di ciascuno e' il figlio diretto di quell'antenato che lo contiene. Cosi'
# non serve conoscere la profondita' esatta del DOM.
_WORK_CARDS_JS = """() => {
  const links = [...document.querySelectorAll('#select-link')];
  if (links.length === 0) return [];
  let lca = links[0];
  while (lca && !links.every(l => lca.contains(l))) lca = lca.parentElement;
  if (!lca) return [];
  return links.map(l => {
    let card = l;
    while (card.parentElement && card.parentElement !== lca) card = card.parentElement;
    const titoli = [...card.querySelectorAll('h1,h2,h3,h4,h5,h6')]
      .map(h => h.innerText.trim())
      .filter(t => t && t.toUpperCase() !== 'SELECT');
    return {
      title: titoli.length ? titoli[0] : '',
      text: card.innerText.replace(/\\s+/g, ' ').trim().toUpperCase()
    };
  });
}"""


# dalla scheda si tiene solo cio' che identifica l'OPERA (titolo, autori,
# publisher) e si scartano i dati della singola REGISTRAZIONE (Song Code,
# Member's Song ID, numero di recordings/artisti collegati): due
# registrazioni doppie della stessa opera differiscono solo su questi ultimi
# (caso reale osservato: PLAYBOOYZ, PN8C4S con 7 recordings e PN8DH1 con 6,
# per il resto identiche) e vanno trattate come la stessa opera.
_SHARES_SPLIT = re.compile(r"TOTAL KNOWN SHARES", re.IGNORECASE)
_PUBLISHERS_RE = re.compile(r"PUBLISHERS \(\d+\)(.*?)(?:RECORDINGS \(|$)", re.IGNORECASE)


def _work_identity_key(card_text: str) -> str:
    """Titolo + autori + publisher di una scheda risultato, senza i dati
    specifici della registrazione."""
    head = _SHARES_SPLIT.split(card_text, maxsplit=1)[0]  # "TITOLO SELECT AUTORI"
    publishers = _PUBLISHERS_RE.search(card_text)
    return f"{head.strip()}|{publishers.group(1).strip() if publishers else ''}"


def _norm_title(titolo: str) -> str:
    """
    Titolo normalizzato per il confronto: senza accenti, maiuscolo, senza
    punteggiatura e con spazi compattati. Cosi' "Killer Instinct - Feb Br
    Remix" e "KILLER INSTINCT (FEB BR REMIX)" risultano lo stesso titolo,
    mentre restano diversi due brani realmente distinti.
    """
    t = unicodedata.normalize("NFKD", titolo or "")
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    t = t.upper().replace("’", "'").replace("`", "'")
    t = re.sub(r"[^A-Z0-9' ]+", " ", t)
    return re.sub(r"\s+", " ", t).strip()


# esiti non numerici di _scegli_opera
TITOLO_ASSENTE = -2   # nessun risultato ha il titolo cercato -> opera non a catalogo
SCELTA_UMANA = -1     # piu' opere omonime ma diverse -> serve un occhio umano


def _scegli_opera(cards: list, titolo_cercato: str) -> tuple[int, str]:
    """
    Decide quale risultato selezionare confrontando il TITOLO dell'opera con
    quello cercato.

    Serve perche' la ricerca titolo di MLC lavora per parole singole e
    restituisce regolarmente opere completamente scorrelate (cercando "The
    FunKing" tornano anche THE DRUMMER, THE PRESSURE, THE CREATOR...): il
    numero di risultati da solo non dice nulla, e nemmeno "un solo risultato"
    significa "risultato giusto".

    Ritorna (indice, motivo), con indice TITOLO_ASSENTE o SCELTA_UMANA nei
    casi non decidibili.
    """
    target = _norm_title(titolo_cercato)
    esatti = [i for i, c in enumerate(cards) if _norm_title(c.get("title", "")) == target]

    if not esatti:
        return TITOLO_ASSENTE, f"nessuno dei {len(cards)} risultati ha il titolo cercato"

    if len(esatti) == 1:
        return esatti[0], f"unico risultato col titolo esatto su {len(cards)}"

    # piu' risultati col titolo giusto: sono la stessa opera registrata piu'
    # volte (differiscono solo per Song Code) oppure opere diverse omonime
    chiavi = {_work_identity_key(cards[i]["text"]) for i in esatti}
    if len(chiavi) == 1:
        return esatti[0], f"{len(esatti)} registrazioni doppie della stessa opera, selezionata la prima"

    return SCELTA_UMANA, f"{len(esatti)} opere diverse con lo stesso titolo (autore/publisher differenti)"


PUBLISHER_DEFAULT_VALUE = "LOO"  # copre "LOOSE CLUB EDITION" a prescindere dal publisher esatto in riga


def _wait_for_any(page: Page, candidates: dict, timeout_ms: int) -> str:
    """
    Attende che compaia UNO qualsiasi degli elementi passati e ne ritorna il
    nome, oppure '' se scade il tempo. Serve quando lo stato della pagina non
    e' prevedibile in anticipo (tipico dopo una pausa manuale: l'utente puo'
    aver lasciato aperto il dialog di conferma oppure aver gia' confermato).
    """
    deadline = time.monotonic() + timeout_ms / 1000
    while time.monotonic() < deadline:
        for nome, locator in candidates.items():
            try:
                if locator.count() > 0 and locator.first.is_visible():
                    return nome
            except Exception:
                pass
        page.wait_for_timeout(250)
    return ""


def _finalize_match(page: Page, note: str) -> tuple[str, str]:
    """
    Chiude il match dopo che l'opera e' stata selezionata.

    Non si assume in che stato sia la pagina: dopo una pausa manuale l'utente
    puo' aver cliccato solo "Select" (dialog di conferma aperto) oppure aver
    gia' fatto Confirm da solo (schermata di successo con "Done"). Cliccare
    "Confirm" alla cieca in quel secondo caso costava 30s di timeout e un
    finto errore (osservato live su una run da 40 righe).
    """
    confirm = page.get_by_role("button", name="Confirm")
    done = page.get_by_role("button", name="Done")

    stato = _wait_for_any(page, {"confirm": confirm, "done": done}, 10000)

    if stato == "":
        return "manual_incomplete", (
            "nessun dialog di conferma ne' schermata di successo dopo la selezione: "
            f"opera probabilmente non selezionata - {note}"
        )

    if stato == "confirm":
        confirm.click()
        # Il portale MLC a volte rifiuta l'invio lato server (HTTP 400
        # "Failed to contact recordings API") SENZA mostrare nulla a schermo:
        # il dialog resta aperto coi pulsanti spariti e "Done" non arriva mai.
        if _wait_for_any(page, {"done": done}, 15000) == "":
            return "submit_failed", f"MLC ha rifiutato l'invio (nessuna conferma dopo il Confirm) - {note}"

    done.click()
    return "matched", note


def stage2_search_work(page: Page, title: str, writer: str, publisher: str, interattivo: bool = False) -> tuple[str, str]:
    """
    Cerca per Titolo + Publisher Name ('LOO' fisso) e, se non basta, ritenta
    con Titolo + Writer Name (cognome autore), piu' selettivo sui titoli
    generici dove il publisher da solo non discrimina.

    In entrambi i tentativi la scelta si fa sul TITOLO dei risultati, mai sul
    loro numero: anche un singolo risultato puo' essere un'opera scorrelata,
    visto che MLC cerca per parole singole.

    Ritorna (esito, nota).
    Esito: 'matched' / 'no_match' / 'ambiguous' / 'submit_failed' /
           'manual_incomplete'
    """
    title_box = page.get_by_test_id("search.0.searchTerm")
    title_box.click()
    title_box.fill(title)

    tentativi = [("Publisher Name", PUBLISHER_DEFAULT_VALUE)]
    if writer:
        tentativi.append(("Writer Name", writer))

    cards, indice, motivo, criterio = [], TITOLO_ASSENTE, "nessun risultato", ""

    for nome_criterio, valore in tentativi:
        criterio = f"Titolo + {nome_criterio} ('{valore}')"
        _set_second_criteria(page, nome_criterio, valore)

        if run_search(page) == "no_results":
            cards, indice, motivo = [], TITOLO_ASSENTE, "nessun risultato"
            continue

        cards = page.evaluate(_WORK_CARDS_JS)
        if not cards:
            indice, motivo = TITOLO_ASSENTE, "nessun risultato"
            continue

        indice, motivo = _scegli_opera(cards, title)
        if indice >= 0:
            break  # trovata: il tentativo successivo non serve

    if indice == TITOLO_ASSENTE:
        # nessun risultato col titolo cercato: l'opera non e' a catalogo su
        # MLC, non c'e' niente da chiedere - si va avanti
        return "no_match", f"{motivo} ({criterio})"

    if indice == SCELTA_UMANA:
        if not interattivo:
            # niente pause a meta' lotto: la riga va nella coda di fine run
            return "ambiguous", f"{motivo} - da scegliere a mano ({criterio})"
        titoli = " | ".join(c.get("title", "?") for c in cards)
        print(f"    -> {motivo} per '{title}': {titoli}")
        input("       Seleziona manualmente l'opera corretta nel browser, poi premi invio qui per continuare...")
        return _finalize_match(page, f"scelta manuale fra {len(cards)} risultati - match su {criterio}")

    page.locator("#select-link").nth(indice).click()
    return _finalize_match(page, f"{motivo} - match su {criterio}")


# ---------------------------------------------------------------------------
# Orchestrazione riga per riga
# ---------------------------------------------------------------------------

def process_row(page: Page, isrc: str, title: str, writer: str, publisher: str, interattivo: bool = False) -> RowResult:
    result = RowResult(isrc=isrc, title=title, writer=writer, publisher=publisher)
    try:
        if not isrc:
            result.status = "error"
            result.note = "ISRC mancante in input"
            reset_search_form(page)
            return result

        stage1_status = stage1_search_isrc(page, isrc)

        if stage1_status == "no_match":
            result.status = "no_match_recording"
            result.note = "Nessuna registrazione trovata per questo ISRC"
            reset_search_form(page)
            return result

        if stage1_status == "already_submitted":
            result.status = "already_submitted"
            result.note = "Gruppo di registrazione gia' inviato/gestito in una sessione precedente"
            reset_search_form(page)
            return result

        if stage1_status == "ambiguous":
            result.status = "ambiguous_recording"
            result.note = "Layout inatteso: nessun controllo di selezione trovato per questo ISRC"
            reset_search_form(page)
            return result

        # Stage 1 ok -> procedi con Stage 2
        stage2_status, stage2_note = stage2_search_work(page, title, writer, publisher, interattivo)
        result.note = stage2_note

        if stage2_status == "matched":
            result.status = "matched"
            reset_search_form(page)
        elif stage2_status == "no_match":
            result.status = "no_match_work"
            abandon_stage2(page)
        elif stage2_status in ("submit_failed", "manual_incomplete"):
            # la UI resta in uno stato intermedio (dialog bloccato o selezione
            # a meta'): solo un reload completo la rimette in sesto
            result.status = stage2_status
            page.reload(wait_until="networkidle")
            _dismiss_cookies(page)
            page.get_by_role("link", name="Matching Tool", exact=True).click()
            page.wait_for_load_state("networkidle")
        else:
            # ambiguous_work: nessun match confermato, quindi i gruppi
            # recording sono ancora selezionati e vanno deselezionati come
            # per no_match_work, altrimenti restano agganciati alla riga dopo
            result.status = "ambiguous_work"
            abandon_stage2(page)

        return result

    except Exception as e:
        result.status = "error"
        result.note = str(e)
        try:
            # reload duro, non un semplice click sulla sidebar: un errore a
            # meta' di un'interazione (es. cambio criterio in Stage 2) puo'
            # lasciare lo stato del form "sporco" per le righe successive
            # (osservato live), un click SPA da solo non lo ripristina
            page.reload(wait_until="networkidle")
            _dismiss_cookies(page)  # il banner puo' ricomparire dopo un reload e bloccare i click
            page.get_by_role("link", name="Matching Tool", exact=True).click()
            page.wait_for_load_state("networkidle")
        except Exception:
            pass
        return result


# ---------------------------------------------------------------------------
# Input / CLI
# ---------------------------------------------------------------------------

def load_input(path: str, sheet, isrc_col: str, title_col: str, writer_col: str, publisher_col: str) -> pd.DataFrame:
    p = Path(path)
    if p.suffix.lower() == ".csv":
        df = pd.read_csv(p, dtype=str)
    else:
        df = pd.read_excel(p, sheet_name=sheet or 0, dtype=str)

    df = df.rename(columns={c: str(c).strip() for c in df.columns})

    required = [isrc_col, title_col]
    for col in required:
        if col not in df.columns:
            raise ValueError(f"Colonna '{col}' non trovata. Colonne disponibili: {list(df.columns)}")

    # writer e publisher sono opzionali: se la colonna manca, si procede senza
    writer_series = df[writer_col] if writer_col in df.columns else ""
    publisher_series = df[publisher_col] if publisher_col in df.columns else ""

    out = pd.DataFrame({
        "isrc": df[isrc_col],
        "title": df[title_col],
        "writer": writer_series,
        "publisher": publisher_series,
    }).fillna("")
    return out


def _color_report(path: str, df: pd.DataFrame) -> None:
    """Colora ogni riga del report Excel in base allo stato, come nel flusso
    manuale: verde se c'e' un match (o e' gia' stato gestito prima), giallo
    se non si trova nulla e serve verifica manuale."""
    wb = load_workbook(path)
    ws = wb.active
    n_cols = len(df.columns)
    for row_idx, status in enumerate(df["status"], start=2):  # riga 1 = header
        if status in GREEN_STATUSES:
            fill = GREEN_FILL
        elif status in YELLOW_STATUSES:
            fill = YELLOW_FILL
        elif status in ORANGE_STATUSES:
            fill = ORANGE_FILL
        else:
            continue
        for col_idx in range(1, n_cols + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill
    wb.save(path)


def _save_report(path: str, results: list) -> pd.DataFrame:
    """
    Scrive il report su disco. Viene richiamato dopo OGNI riga, non solo a
    fine run: su lotti lunghi (40+ righe, con pause manuali in mezzo) un
    Ctrl+C o un crash altrimenti butterebbe via tutto il lavoro gia' fatto.
    """
    out_df = pd.DataFrame([r.__dict__ for r in results])
    if out_df.empty:
        return out_df
    out_df.to_excel(path, index=False)
    _color_report(path, out_df)
    return out_df


def main():
    parser = argparse.ArgumentParser(description="Automazione matching su The MLC Matching Tool")
    parser.add_argument("input_file", help="Percorso Excel/CSV con il catalogo da matchare")
    parser.add_argument("--sheet", default=None, help="Nome sheet Excel (default: primo)")
    parser.add_argument("--isrc-col", default="ISRC Code")
    parser.add_argument("--title-col", default="Track Title")
    parser.add_argument("--writer-col", default="Surname", help="Colonna cognome autore, usata come fallback allo Stage 2 se Publisher Name ('LOO') non trova un risultato univoco")
    parser.add_argument("--publisher-col", default="Publisher Name", help="Non usato per la ricerca (si usa il valore fisso 'LOO'), tenuto solo per riferimento nel report")
    parser.add_argument("--output", default="mlc_match_results.xlsx")
    parser.add_argument("--headless", action="store_true", help="Esegui senza finestra browser visibile")
    parser.add_argument("--skip-ambigui", action="store_true", help="Salta anche la coda manuale di fine run: le righe ambigue restano segnate ambiguous_work nel report, da sistemare in un secondo momento")
    args = parser.parse_args()

    email = os.environ.get("MLC_EMAIL")
    password = os.environ.get("MLC_PASSWORD")
    if not email or not password:
        sys.exit("Imposta le variabili d'ambiente MLC_EMAIL e MLC_PASSWORD prima di avviare lo script.")

    df = load_input(args.input_file, args.sheet, args.isrc_col, args.title_col, args.writer_col, args.publisher_col)
    print(f"Righe da processare: {len(df)}")

    results = []
    interrotto = False
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=args.headless)
            page = browser.new_page()
            login(page, email, password)

            # Prima passata: nessuna pausa, cosi' il lotto gira da solo.
            # Le righe che richiedono un occhio umano vengono messe da parte.
            for i, row in df.iterrows():
                print(f"[{i + 1}/{len(df)}] ISRC={row['isrc']} Title={row['title']} Writer={row['writer']}")
                r = process_row(page, row["isrc"], row["title"], row["writer"], row["publisher"], interattivo=False)
                print(f"  -> {r.status} ({r.note})")
                results.append(r)
                _save_report(args.output, results)  # salvataggio dopo ogni riga

            # Seconda passata: le sole righe ambigue, tutte in coda, quando il
            # grosso del lavoro e' gia' fatto e salvato
            da_rivedere = [r for r in results if r.status == "ambiguous_work"]
            if da_rivedere and not args.skip_ambigui:
                print(f"\n{'=' * 60}")
                print(f"Restano {len(da_rivedere)} righe da scegliere a mano.")
                print(f"{'=' * 60}")
                for n, r in enumerate(da_rivedere, 1):
                    print(f"\n[manuale {n}/{len(da_rivedere)}] ISRC={r.isrc} Title={r.title}")
                    nuovo = process_row(page, r.isrc, r.title, r.writer, r.publisher, interattivo=True)
                    print(f"  -> {nuovo.status} ({nuovo.note})")
                    r.status, r.note = nuovo.status, nuovo.note
                    _save_report(args.output, results)
            elif da_rivedere:
                print(f"\n{len(da_rivedere)} righe ambigue lasciate da rivedere a mano (--skip-ambigui).")

            browser.close()
    except KeyboardInterrupt:
        interrotto = True
        print("\nInterrotto: i risultati delle righe gia' processate sono stati salvati.")

    out_df = _save_report(args.output, results)
    if out_df.empty:
        print("\nNessuna riga processata, report non scritto.")
        return

    print(f"\nReport salvato in: {args.output}")
    if interrotto:
        print(f"Righe processate: {len(results)} su {len(df)}")
    print(out_df["status"].value_counts())


if __name__ == "__main__":
    main()
